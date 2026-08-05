import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, os

def generate():
    wb = openpyxl.Workbook()

    # ─── STYLES ────────────────────────────────────────────────────────────────
    title_font   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    hdr_font     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font    = Font(name="Calibri", size=10, bold=True, color="1F3864")
    reg_font     = Font(name="Calibri", size=10, color="000000")
    crit_font    = Font(name="Calibri", size=10, bold=True, color="9C0006")
    high_font    = Font(name="Calibri", size=10, bold=True, color="843C00")
    med_font     = Font(name="Calibri", size=10, bold=True, color="7D6608")
    low_font     = Font(name="Calibri", size=10, bold=True, color="004085")

    crit_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    high_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    med_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill  = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    hdr_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    sub_fill  = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    thin = Border(
        left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),  bottom=Side(style='thin', color='BFBFBF'))

    def hdr_row(ws, row, cols, values, fill):
        for c, v in zip(cols, values):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = hdr_font; cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_row(ws, row, cols, values):
        for c, v in zip(cols, values):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = reg_font; cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    def title_cell(ws, addr, text, fill):
        cell = ws[addr]; cell.value = text
        cell.font = title_font; cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def auto_width(ws, max_w=55):
        for col in ws.columns:
            l = get_column_letter(col[0].column)
            mx = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[l].width = min(max(mx + 2, 12), max_w)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Security Findings
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Security Findings"
    ws1.merge_cells("A1:I2")
    title_cell(ws1, "A1", "FitTrack AI — Security Findings (SAST + Manual Review)", hdr_fill)
    ws1.row_dimensions[1].height = 32
    ws1.row_dimensions[2].height = 32

    cols1 = list(range(1, 10))
    hdr_row(ws1, 3, cols1,
        ["ID","Severity","Type","File Path","Endpoint / Screen",
         "Description","Exploitation Scenario","Impact","Recommended Fix"], sub_fill)
    ws1.row_dimensions[3].height = 28

    findings = [
        ("C-01","Critical","Insecure Credential Storage (CWE-312)",
         "backend/PreferenceManager.java L254",
         "N/A — local storage",
         "User password stored as plaintext in SharedPreferences XML file.",
         "Attacker with ADB/root access reads FitTrackPrefs.xml and extracts passwords for all registered users.",
         "Full credential compromise; credential stuffing on external services.",
         "Remove KEY_USER_PASSWORD entirely. Use Firebase Auth session tokens. Implement EncryptedSharedPreferences."),

        ("C-02","Critical","Hardcoded API Keys / Secret Leakage (CWE-798)",
         "app/src/main/google-services.json",
         "N/A — config file",
         "Firebase API key, project ID, and OAuth client ID committed to Git repository and visible to all contributors.",
         "Clone repo → extract api_key → issue direct Firestore REST API calls bypassing app logic.",
         "Unauthorized data access, API quota abuse, potential account enumeration.",
         "Add google-services.json to .gitignore. Rotate all Firebase keys. Store in GitHub Secrets for CI."),

        ("C-03","Critical","Missing Server-Side Authorization (CWE-285)",
         "Firebase Console (Firestore Security Rules)",
         "Firestore: users/{uid}",
         "No confirmed Firestore Security Rules. Any authenticated Firebase user can read/write any document.",
         "Authenticated user substitutes victim UID in Firestore REST call → reads health/diet/weight data.",
         "Full user health data exposure. GDPR violation risk.",
         "Deploy: allow read,write: if request.auth.uid == userId; in Firestore Security Rules."),

        ("H-01","High","Sensitive Data in localStorage (CWE-312)",
         "web-react/src/App.jsx",
         "React Web — all views",
         "Health data (stress level, mood, water count, calories) stored in browser localStorage in plaintext.",
         "XSS payload reads localStorage and exfiltrates health data to attacker-controlled server.",
         "Health data exposure; HIPAA/GDPR noncompliance if deployed publicly.",
         "Store sensitive health data server-side in Firestore. Use sessionStorage for ephemeral UI state."),

        ("H-02","High","Debug Auth Fallback in Production Build (CWE-287)",
         "frontend/LoginActivity.java",
         "LoginActivity — Google Sign-In",
         "Debug fallback path creates mock user profile without actual token validation if Google Sign-In fails.",
         "If debug APK is distributed, attacker triggers Google Sign-In failure to gain mock authenticated session.",
         "Unauthorized access to authenticated app screens.",
         "Wrap all debug-only code in BuildConfig.DEBUG checks. Add ProGuard rule to strip debug blocks."),

        ("H-03","High","Intent Extra Injection — No Bounds Validation (CWE-20)",
         "frontend/YogaHomeActivity.java, ActiveWorkoutActivity.java",
         "YogaHomeActivity, ActiveWorkoutActivity",
         "stress_level, duration_seconds, reps read from Intent without range or type validation.",
         "Malicious app sends crafted Intent with negative duration → timer overflow → app crash or logic bypass.",
         "App crash (DoS), logic bypass in workout session tracking.",
         "Validate all Intent extras. Set android:exported=false for all internal Activities in Manifest."),

        ("M-01","Medium","Unvalidated External URL in Glide (CWE-601)",
         "frontend/YogaHomeActivity.java, DietHomeActivity.java, ProfileActivity.java",
         "Multiple screens loading images",
         "Image URLs from Intent extras and Firestore passed directly to Glide without scheme validation.",
         "Attacker sets profileImage field to file:///data/... URL → Glide reads local sensitive file.",
         "Local file disclosure via Glide image loader.",
         "Validate URL scheme before Glide.load(). Only allow https:// and content:// URIs."),

        ("M-02","Medium","Sensitive Data in LogCat Logs (CWE-532)",
         "frontend/LoginActivity.java",
         "LoginActivity",
         "Authentication errors logged to LogCat with email addresses in debug builds.",
         "Developer with ADB access or shared device reads LogCat and extracts user emails.",
         "User PII exposure during development and testing.",
         "Guard all auth-related Log statements with if(BuildConfig.DEBUG). Enable ProGuard log stripping."),

        ("M-03","Medium","No Rate Limiting on Login (CWE-307)",
         "web-react/src/App.jsx",
         "Web Login form",
         "Login form submits to Firebase Auth with no client-side throttle, lockout, or CAPTCHA.",
         "Automated script submits thousands of credential combinations against Firebase Auth endpoint.",
         "Account brute-force, account enumeration via error messages.",
         "Add exponential backoff after failed attempts. Integrate reCAPTCHA v3. Enable Firebase App Check."),

        ("M-04","Medium","Unencrypted SharedPreferences for Health Data (CWE-312)",
         "backend/PreferenceManager.java",
         "N/A — local storage",
         "Health conditions, weight history, meal plans stored as plaintext XML in SharedPreferences.",
         "Rooted device or forensic tool reads SharedPreferences XML and extracts complete health profile.",
         "Sensitive health data exposure on rooted or seized devices.",
         "Migrate to EncryptedSharedPreferences (androidx.security:security-crypto)."),

        ("M-05","Medium","Missing android:exported=false (CWE-926)",
         "app/src/main/AndroidManifest.xml",
         "All internal Activities",
         "Internal Activities may be inadvertently exported allowing external app invocation with crafted Intents.",
         "External malicious app launches ActiveWorkoutActivity with negative reps → crash or data corruption.",
         "Unauthorized Activity invocation, data corruption, UI spoofing.",
         "Add android:exported=false to all non-launcher Activities in AndroidManifest.xml."),

        ("L-01","Low","No Certificate Pinning (CWE-319)",
         "app/src/main/java/... (network layer)",
         "All network calls",
         "No certificate pinning implemented on Firebase SDK calls or Unsplash CDN image loads.",
         "MITM on a compromised network substitutes Firebase responses or injects malicious image content.",
         "Data interception in hostile network environments.",
         "Implement OkHttp CertificatePinner for production Firebase API hostnames."),

        ("L-02","Low","Missing Network Security Config (CWE-319)",
         "app/src/main/res/xml/ (missing)",
         "All network calls",
         "No network_security_config.xml defined. Cleartext traffic may be permitted on Android < 9.",
         "Older Android device sends HTTP request in cleartext → network observer captures data.",
         "Data interception on older Android versions.",
         "Add res/xml/network_security_config.xml with cleartextTrafficPermitted=false."),

        ("L-03","Low","Unguarded Float.parseFloat() (CWE-390)",
         "backend/PreferenceManager.java L234",
         "N/A — local data",
         "Weight history parsed with Float.parseFloat() without try-catch block.",
         "Corrupted SharedPreferences data triggers uncaught NumberFormatException → app crash.",
         "App crash (DoS) on corrupt stored data.",
         "Wrap Float.parseFloat() in try-catch NumberFormatException with graceful fallback value."),
    ]

    sev_map = {"Critical":(crit_font,crit_fill),"High":(high_font,high_fill),
               "Medium":(med_font,med_fill),"Low":(low_font,low_fill)}

    for i, f in enumerate(findings, start=4):
        data_row(ws1, i, cols1, f)
        sev_cell = ws1.cell(row=i, column=2)
        fn, ff = sev_map.get(f[1], (reg_font, None))
        sev_cell.font = fn
        if ff: sev_cell.fill = ff
        sev_cell.alignment = Alignment(horizontal="center", vertical="top")
        ws1.row_dimensions[i].height = 65

    auto_width(ws1)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Endpoint Inventory
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Endpoint Inventory")
    ws2.merge_cells("A1:F2")
    title_cell(ws2, "A1", "FitTrack AI — Full Endpoint & Screen Inventory", hdr_fill)
    ws2.row_dimensions[1].height = 32; ws2.row_dimensions[2].height = 32

    cols2 = list(range(1, 7))
    hdr_row(ws2, 3, cols2,
        ["Screen / Endpoint","Method / Launch","Auth Required","Expected Role","File Path","Notes"], sub_fill)

    endpoints = [
        ("SplashActivity","Intent","No","Public","frontend/SplashActivity.java","Entry point — auto redirects"),
        ("LoginActivity","Intent","No","Public","frontend/LoginActivity.java","Email+Password & Google Sign-In"),
        ("SignUpActivity","Intent","No","Public","frontend/SignUpActivity.java","Creates Firestore user document"),
        ("OnboardingActivity","Intent","Post-Signup","Registered User","frontend/OnboardingActivity.java","Onboarding wizard"),
        ("DashboardActivity","Intent","Yes","Authenticated","frontend/DashboardActivity.java","Main hub — all stats"),
        ("DietHomeActivity","Intent","Yes","Authenticated","frontend/DietHomeActivity.java","7-day rotational meals"),
        ("WorkoutHomeActivity","Intent","Yes","Authenticated","frontend/WorkoutHomeActivity.java","Workout catalog & filtering"),
        ("YogaHomeActivity","Intent","Yes","Authenticated","frontend/YogaHomeActivity.java","Stress-based Pranayama"),
        ("StressMapperActivity","Intent","Yes","Authenticated","frontend/StressMapperActivity.java","Stress slider + mood"),
        ("ProfileActivity","Intent","Yes","Authenticated","frontend/ProfileActivity.java","View profile & stats"),
        ("EditProfileActivity","Intent","Yes","Authenticated","frontend/EditProfileActivity.java","Edit name/height/avatar"),
        ("WaterTrackerActivity","Intent","Yes","Authenticated","frontend/WaterTrackerActivity.java","Hydration logging"),
        ("ProgressActivity","Intent","Yes","Authenticated","frontend/ProgressActivity.java","Weight chart & history"),
        ("RegionalDietActivity","Intent","Yes","Authenticated","frontend/RegionalDietActivity.java","South/North Indian diet"),
        ("ActiveWorkoutActivity","Intent","Yes","Authenticated","frontend/ActiveWorkoutActivity.java","Countdown timer + sets"),
        ("PantryGeneratorActivity","Intent","Yes","Authenticated","frontend/PantryGeneratorActivity.java","AI recipe from ingredients"),
        ("SettingsActivity","Intent","Yes","Authenticated","frontend/SettingsActivity.java","Dark mode, notifications"),
        ("MicroWorkoutActivity","Intent","Yes","Authenticated","frontend/MicroWorkoutActivity.java","Short workout sessions"),
        ("Firestore: users/{uid} GET","Firebase REST","Yes","Owner UID only","LoginActivity.java","⚠️ Needs Security Rules"),
        ("Firestore: users/{uid} SET","Firebase REST","Yes","Owner UID only","SignUpActivity.java, EditProfileActivity.java","⚠️ Needs Security Rules"),
        ("Firebase Auth: signIn","Firebase REST","No","Public","LoginActivity.java, App.jsx","No rate limiting"),
        ("Firebase Auth: createUser","Firebase REST","No","Public","SignUpActivity.java, App.jsx","No rate limiting"),
        ("Firebase Auth: googlePopup","Firebase REST","No","Public","LoginActivity.java, App.jsx","⚠️ Debug fallback"),
        ("Firebase Auth: resetPassword","Firebase REST","No","Public","LoginActivity.java, App.jsx","Email enumeration risk"),
        ("Unsplash CDN (images)","HTTPS GET","No","Public","DietData.java, App.jsx","External CDN — no pinning"),
        ("Web: /login","React Route","No","Public","web-react/src/App.jsx","Auth form — no CAPTCHA"),
        ("Web: /dashboard","React State","Yes","Authenticated","web-react/src/App.jsx","All dashboard views"),
        ("Web: /diet","React State","Yes","Authenticated","web-react/src/App.jsx","Meal plan views"),
        ("Web: /workout","React State","Yes","Authenticated","web-react/src/App.jsx","Workout + Yoga cards"),
        ("Web: /profile","React State","Yes","Authenticated","web-react/src/App.jsx","Edit profile + avatar"),
    ]

    for i, ep in enumerate(endpoints, start=4):
        data_row(ws2, i, cols2, ep)
        ws2.row_dimensions[i].height = 22

    auto_width(ws2)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Dependency Vulnerabilities
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Dependency Vulnerabilities")
    ws3.merge_cells("A1:G2")
    title_cell(ws3, "A1", "FitTrack AI — Dependency Risk Report", hdr_fill)
    ws3.row_dimensions[1].height = 32; ws3.row_dimensions[2].height = 32

    cols3 = list(range(1, 8))
    hdr_row(ws3, 3, cols3,
        ["Package","Platform","Current Usage","Risk Level","CVE / Notes","Remediation","Priority"], sub_fill)

    deps = [
        ("com.google.firebase:firebase-auth","Android","Firebase Authentication","Critical","API key exposed in repo","Rotate keys. Add google-services.json to .gitignore","P0"),
        ("com.google.firebase:firebase-firestore","Android","Firestore database access","Critical","No confirmed Security Rules → all user data exposed","Deploy Firestore Security Rules immediately","P0"),
        ("androidx.security:security-crypto","Android","NOT USED — Missing dependency","Critical","SharedPreferences stored in plaintext","Add security-crypto:1.1.0-alpha06. Migrate to EncryptedSharedPreferences","P0"),
        ("com.github.bumptech.glide:glide","Android","External image loading","Medium","External URLs loaded without scheme validation","Validate URL scheme before load(). Block file:// URIs","P1"),
        ("com.google.android.material","Android","UI components","Low","Ensure latest stable version","Run ./gradlew dependencyUpdates","P3"),
        ("firebase (JS SDK)","Web (React)","Auth + Firestore","Critical","Firebase config committed to repo in firebase.js","Rotate keys. Move to env variables. Do not commit secrets.","P0"),
        ("vite","Web (React)","Build tool / dev server","Medium","Historical XSS in dev server in versions < 4.x","Ensure latest vite patch version","P2"),
        ("react / react-dom","Web (React)","UI framework","Low","Keep updated for security patches","npm update react react-dom","P3"),
        ("recharts","Web (React)","Chart rendering","Low","Low attack surface — UI only","Keep updated","P3"),
        ("lucide-react","Web (React)","Icon library","Low","UI icons only — low risk","Keep updated","P3"),
    ]

    for i, d in enumerate(deps, start=4):
        data_row(ws3, i, cols3, d)
        risk_cell = ws3.cell(row=i, column=4)
        fn, ff = sev_map.get(d[3], (reg_font, None))
        risk_cell.font = fn
        if ff: risk_cell.fill = ff
        risk_cell.alignment = Alignment(horizontal="center", vertical="top")
        ws3.row_dimensions[i].height = 35

    auto_width(ws3)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Risk Summary
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Risk Summary")
    ws4.merge_cells("A1:D2")
    title_cell(ws4, "A1", "FitTrack AI — Executive Risk Summary Dashboard", hdr_fill)
    ws4.row_dimensions[1].height = 32; ws4.row_dimensions[2].height = 32

    # KPI block
    kpis = [
        ("A4", "Assessment Date", datetime.datetime.now().strftime("%Y-%m-%d")),
        ("A5", "Total Findings", "14"),
        ("A6", "Critical", "3"),
        ("A7", "High", "3"),
        ("A8", "Medium", "5"),
        ("A9", "Low", "3"),
        ("A11","Overall Security Score", "54 / 100"),
        ("A12","OWASP M-Top10 Compliant", "4 / 10 categories passing"),
    ]
    for addr, label, val in kpis:
        lc = ws4[addr]; vc = ws4[addr.replace("A","B")]
        lc.value = label; lc.font = bold_font
        vc.value = val; vc.font = reg_font
        if label == "Critical": vc.font = crit_font; vc.fill = crit_fill
        if label == "High": vc.font = high_font; vc.fill = high_fill
        if label == "Medium": vc.font = med_font; vc.fill = med_fill
        if label == "Low": vc.font = low_font; vc.fill = low_fill

    # Priority table
    hdr_row(ws4, 14, [1,2,3,4],
        ["Priority","Finding ID","Action Required","Due Sprint"], sub_fill)

    prios = [
        ("P0 — Immediate","C-01","Remove plaintext password from SharedPreferences","This Week"),
        ("P0 — Immediate","C-02","Rotate Firebase API keys. Add google-services.json to .gitignore","This Week"),
        ("P0 — Immediate","C-03","Deploy Firestore Security Rules before any public release","This Week"),
        ("P1 — This Sprint","H-01","Migrate health data from localStorage to Firestore","Sprint 1"),
        ("P1 — This Sprint","H-02","Strip debug fallback code from release builds","Sprint 1"),
        ("P1 — This Sprint","H-03","Validate all Intent extras with bounds check","Sprint 1"),
        ("P1 — This Sprint","M-04","Migrate to EncryptedSharedPreferences","Sprint 1"),
        ("P2 — Next Sprint","M-01","Validate image URLs before Glide.load()","Sprint 2"),
        ("P2 — Next Sprint","M-02","Remove sensitive LogCat logging from auth flows","Sprint 2"),
        ("P2 — Next Sprint","M-03","Add reCAPTCHA v3 and login rate limiting to web app","Sprint 2"),
        ("P2 — Next Sprint","M-05","Set android:exported=false for all internal Activities","Sprint 2"),
        ("P3 — Backlog","L-01","Implement OkHttp certificate pinning","Backlog"),
        ("P3 — Backlog","L-02","Add network_security_config.xml","Backlog"),
        ("P3 — Backlog","L-03","Add try-catch around Float.parseFloat() in weight history","Backlog"),
    ]

    for i, p in enumerate(prios, start=15):
        data_row(ws4, i, [1,2,3,4], p)
        ws4.row_dimensions[i].height = 22

    auto_width(ws4)

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Vulnerability Test Results", "findings.xlsx")
    wb.save(out)
    print(f"[OK] Security findings Excel report generated: {out}")

if __name__ == "__main__":
    generate()
